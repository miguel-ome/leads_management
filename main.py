import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import os

ARQUIVO_EXCEL = "clientes.xlsx"

CAMPOS = [
    "ID", "Nome", "Cidade", "Estado", "Celular", "Dt. Nasc.", "Email", "Salário", "Categoria", "Objetivo", "Motivação", "T. de Compra", "Orçamento", "F. Pagamento" ,"Observações"
]

larguras_colunas = {
    "ID": 5,
    "Nome": 15,
    "Cidade": 12,
    "Estado": 6,
    "Celular": 12,
    "Dt. Nasc.": 11,
    "Email": 18,
    "Salário": 10,
    "Categoria": 11,
    "Objetivo": 14,
    "Motivação": 16,
    "T. de Compra": 12,
    "Orçamento": 15,
    "F. Pagamento": 15,
    "Observações": 20
}


# Paleta de cores
COR_BG = "#465F87"
COR_FUNDO = "#324562"
COR_TITULO = "#c2e4e4"
COR_BOTAO = "#d2f4f4"

class SistemaClientes:
    def __init__(self, root):
        self.root = root
        self.root.title("Sistema de Clientes")
        self.root.geometry("1400x600")
        self.root.configure(bg=COR_BG)
        self.root.state('zoomed')

        self.dados = []
        self.entries_filtro = []

        self.criar_excel_se_nao_existir()
        self.criar_widgets()
        self.carregar_dados()

    def criar_excel_se_nao_existir(self):
        if not os.path.exists(ARQUIVO_EXCEL):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(CAMPOS)
            wb.save(ARQUIVO_EXCEL)
            
    def criar_widgets(self):
        self.frame_principal = tk.Frame(self.root, bg=COR_BG, bd=0, highlightthickness=0)
        self.frame_principal.pack(fill=tk.BOTH, expand=True)

        # Parte superior: filtros + botão Filtrar
        self.frame_filtros = tk.Frame(self.frame_principal, bg=COR_BG)
        self.frame_filtros.pack(fill=tk.X, padx=10, pady=(10, 5))

        for i, campo in enumerate(CAMPOS):
            lbl = tk.Label(self.frame_filtros, text=campo, bg=COR_FUNDO, fg='white')
            lbl.grid(row=0, column=i, sticky="ew")

            if campo == "Salário":
                frame_salario = tk.Frame(self.frame_filtros, bg=COR_BG)
                frame_salario.grid(row=1, column=i, sticky="ew", padx=1, pady=1)
                
                var_operador = tk.StringVar(value="=")
                operador = ttk.Combobox(frame_salario, textvariable=var_operador, width=2, state="readonly")
                operador['values'] = ("=", ">", "<", ">=", "<=")
                operador.grid(row=0, column=0)

                entry_valor = tk.Entry(frame_salario, width=15)
                entry_valor.grid(row=0, column=1)

                self.entries_filtro.append((var_operador, entry_valor))
            else:
                entry = tk.Entry(self.frame_filtros, width=larguras_colunas[campo])
                entry.grid(row=1, column=i, sticky="ew", padx=1, pady=1)
                self.entries_filtro.append(entry)

            self.frame_filtros.grid_columnconfigure(i, weight=1)


        botao_filtro = tk.Button(self.frame_filtros, text="Filtrar", command=self.filtrar, bg=COR_FUNDO, fg='white')
        botao_filtro.grid(row=1, column=len(CAMPOS), padx=(10, 0))

        # Frame da Tabela
        frame_tabela = tk.Frame(self.frame_principal, bg=COR_BG)
        frame_tabela.pack(fill=tk.BOTH, expand=True, padx=10, pady=(5, 10))

        # Scrollbars
        scrollbar_y = ttk.Scrollbar(frame_tabela, orient="vertical")
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

        scrollbar_x = ttk.Scrollbar(frame_tabela, orient="horizontal")
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)

        # Tabela (Treeview)
        self.tabela = ttk.Treeview(
            frame_tabela,
            columns=CAMPOS,
            show="headings",
            yscrollcommand=scrollbar_y.set,
            xscrollcommand=scrollbar_x.set
        )
        for campo in CAMPOS:
            self.tabela.heading(campo, text=campo)
            self.tabela.column(campo, width=larguras_colunas[campo] * 10, anchor="w")

        self.tabela.pack(fill=tk.BOTH, expand=True)

        scrollbar_y.config(command=self.tabela.yview)
        scrollbar_x.config(command=self.tabela.xview)

        # Botões CRUD
        frame_botoes = tk.Frame(self.root, bg=COR_BG)
        frame_botoes.pack(pady=10, padx=20)

        tk.Button(frame_botoes, text="Criar", command=self.criar_cliente, bg=COR_FUNDO, width=12, fg='white', bd=0, highlightthickness=0).pack(side=tk.LEFT, padx=10)
        tk.Button(frame_botoes, text="Editar", command=self.editar_cliente, bg=COR_FUNDO, width=12, fg='white', bd=0, highlightthickness=0).pack(side=tk.LEFT, padx=10)
        tk.Button(frame_botoes, text="Excluir", command=self.excluir_cliente, bg=COR_FUNDO, width=12, fg='white', bd=0, highlightthickness=0).pack(side=tk.LEFT, padx=10)

    def carregar_dados(self):
        self.dados.clear()
        wb = openpyxl.load_workbook(ARQUIVO_EXCEL)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            self.dados.append(row)
        wb.close()
        self.atualizar_tabela(self.dados)

    def atualizar_tabela(self, dados):
        for item in self.tabela.get_children():
            self.tabela.delete(item)
        for row in dados:
            # Substitui valores None por string vazia antes de inserir
            linha_formatada = [cell if cell is not None else "" for cell in row]
            self.tabela.insert("", tk.END, values=linha_formatada)

    def filtrar(self):
        resultado = []

        for row in self.dados:
            manter = True

            for i in range(len(CAMPOS)):
                campo = CAMPOS[i]
                dado = str(row[i]).lower()

                if campo == "Salário":
                    operador, entry = self.entries_filtro[i]
                    op = operador.get().strip()
                    val = entry.get().strip()

                    if val:
                        try:
                            dado_float = float(row[i])
                            val_float = float(val)

                            if op == "=" and not (dado_float == val_float):
                                manter = False
                            elif op == ">" and not (dado_float > val_float):
                                manter = False
                            elif op == "<" and not (dado_float < val_float):
                                manter = False
                            elif op == ">=" and not (dado_float >= val_float):
                                manter = False
                            elif op == "<=" and not (dado_float <= val_float):
                                manter = False
                        except ValueError:
                            manter = False
                else:
                    filtro = self.entries_filtro[i].get().lower()
                    if filtro and filtro not in dado:
                        manter = False

            if manter:
                resultado.append(row)

        self.atualizar_tabela(resultado)


    def criar_cliente(self):
        self.abrir_janela_edicao("Criar")

    def editar_cliente(self):
        selecionado = self.tabela.focus()
        if not selecionado:
            messagebox.showwarning("Atenção", "Selecione um cliente para editar.")
            return
        valores = self.tabela.item(selecionado)["values"]
        self.abrir_janela_edicao("Editar", valores)

    def excluir_cliente(self):
        selecionado = self.tabela.focus()
        if not selecionado:
            messagebox.showwarning("Atenção", "Selecione um cliente para excluir.")
            return
        valores = self.tabela.item(selecionado)["values"]
        if messagebox.askyesno("Confirmação", f"Deseja excluir o cliente {valores[1]}?"):
            wb = openpyxl.load_workbook(ARQUIVO_EXCEL)
            ws = wb.active
            for row in ws.iter_rows(min_row=2):
                if row[0].value == valores[0]:
                    ws.delete_rows(row[0].row)
                    break
            wb.save(ARQUIVO_EXCEL)
            self.carregar_dados()
            
    def abrir_janela_edicao(self, modo, dados=None):
        janela = tk.Toplevel(self.root)
        janela.title("Cadastrar Cliente" if modo == "Criar" else "Atualizar Cadastro")
        janela.configure(bg=COR_BG)

        largura_janela = 600
        altura_janela = 500

        janela.update_idletasks()
        x = (janela.winfo_screenwidth() // 2) - (largura_janela // 2)
        y = (janela.winfo_screenheight() // 2) - (altura_janela // 2)
        janela.geometry(f"{largura_janela}x{altura_janela}+{x}+{y}")
        janela.resizable(False, False)

        titulo = tk.Label(janela, text=janela.title(), font=("Arial", 14, "bold"), fg="white", bg=COR_BG)
        titulo.pack(pady=10)

        # --- Scroll Setup ---
        container = tk.Frame(janela, bg=COR_BG)
        container.pack(fill="both", expand=True)

        canvas = tk.Canvas(container, bg=COR_BG, highlightthickness=0)
        scrollbar = tk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg=COR_BG)
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        scroll_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        frame_central = tk.Frame(scroll_frame, bg=COR_BG)
        frame_central.pack(padx=50, pady=10)  # <-- padding para centralizar mais

        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        entradas = []
        for i, campo in enumerate(CAMPOS[1:]):
            if campo == "Observações":
                lbl = tk.Label(frame_central, text=f"{campo :}", bg=COR_BG, fg="white")
                lbl.grid(row=i * 2, column=0, columnspan=2, sticky="w", padx=10, pady=(10, 2))

                text = tk.Text(frame_central, width=65, height=5)
                text.grid(row=i * 2 + 1, column=0, columnspan=2, sticky="w" ,padx=10, pady=(0, 10))
                if dados:
                    text.insert("1.0", dados[i + 1])
                entradas.append(text)
            else:
                linha = i // 2
                coluna = i % 2

                lbl = tk.Label(frame_central, text=f"{campo :}", bg=COR_BG, fg="white")
                lbl.grid(row=linha * 2, column=coluna, sticky="w", padx=10, pady=(10, 2))

                entry = tk.Entry(frame_central, width=30)
                entry.grid(row=linha * 2 + 1, column=coluna, sticky="w" ,padx=10, pady=(0, 10))

                if dados:
                    entry.insert(0, dados[i + 1])
                entradas.append(entry)


        def salvar():
            valores = []
            for e in entradas:
                if isinstance(e, tk.Text):
                    valores.append(e.get("1.0", "end").strip())
                else:
                    valores.append(e.get().strip())

            wb = openpyxl.load_workbook(ARQUIVO_EXCEL)
            ws = wb.active

            if modo == "Criar":
                novo_id = ws.max_row
                ws.append([novo_id] + valores)
            else:
                for row in ws.iter_rows(min_row=2):
                    if row[0].value == dados[0]:
                        for i, val in enumerate(valores):
                            row[i + 1].value = val
                        break
            wb.save(ARQUIVO_EXCEL)
            janela.destroy()
            self.carregar_dados()

        botao_salvar = tk.Button(
            janela,
            text="Salvar",
            command=salvar,
            bg="white",
            fg="black",
            width=15
        )
        botao_salvar.pack(pady=10)

# Executar
if __name__ == "__main__":
    root = tk.Tk()
    app = SistemaClientes(root)
    root.mainloop()
