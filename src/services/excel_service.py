import os

import openpyxl

from config import FIELDS_TABLE_LEADS

class Excel_Services:
  path_excel = os.path.join(os.path.dirname(__file__), '..', 'database')

  @staticmethod
  def create_database_file(name=""):
    if not name: 
      raise("Campo name obrigatório")
    
    path_file = os.path.join(Excel_Services.path_excel, name)
    
    if not os.path.exists(path_file):
      wb = openpyxl.Workbook()
      ws = wb.active
      ws.append(FIELDS_TABLE_LEADS)
      wb.save(path_file)

  @staticmethod
  def load_data(name=""):
    if not name: 
      raise("Campo name obrigatório")

    path_file = os.path.join(Excel_Services.path_excel, name)

    wb = openpyxl.load_workbook(path_file)
    ws = wb.active
    return [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]
  
  @staticmethod
  def validate_exist_DB():
    arquivos = [f for f in os.listdir(Excel_Services.path_excel) if os.path.isfile(os.path.join(Excel_Services.path_excel, f))]
    if arquivos:
      return True 
    else : 
      return False
    